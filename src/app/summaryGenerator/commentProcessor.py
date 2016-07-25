# Description:
# Date: 2016-01-17

import logging as lg
import openpyxl
import os
import ConfigParser as mParser


class CommentProcessor(object):
    def __init__(self):
        # excel content info, includes: head -> table headers,
        # body -> table content, column -> set of column options
        # body structure: [child_ID] => {head:[values]}
        self.config = {'special_case_file':"", 'preliminary': True}
        self.content = {}
        self.switch = {} # output info
        self.its_to_remove = []

    def open_comment_file(self, filename):
        if filename == None:
            return

        self.config['special_case_file'] = filename
        lg.debug("open comment file: " + self.config['special_case_file'])
        workbook = openpyxl.load_workbook(self.config['special_case_file'])
        sheet = workbook.get_sheet_by_name("Special Cases")
        excel_head = [x.value for x in sheet.rows[0] if x.value is not None]

        # get content body list
        # for each record is a dictionary
        excel_body = []
        for rnum in range(2, sheet.max_row+1):
            # sometimes the max_row number is not accurate, so I add an addition check
            if sheet.cell(row=rnum, column=2).value is None:
                break

            if sheet.cell(row=rnum, column=1).value is not None:
                child_dict = {}
                for i, item in enumerate(excel_head):
                    if sheet.cell(row=rnum, column=i+1).value is None:
                        child_dict[item] = [" "]
                    else:
                        if item == "Recording":
                            child_dict[item] = [str(sheet.cell(row=rnum, column=i+1).value)]
                        elif item == "Recording Dates":
                            child_dict[item] = [sheet.cell(row=rnum, column=i+1).value.strftime("%Y-%m-%d")]
                        else:
                            child_dict[item] = [sheet.cell(row=rnum, column=i+1).value]
                excel_body.append(child_dict)
            else:
                # next record for the same child
                for i, item in enumerate(excel_head):
                    if sheet.cell(row=rnum, column=i+1).value is not None:
                        if item == "Recording":
                            excel_body[-1][item].append(str(sheet.cell(row=rnum,
                                                                       column=i+1).value))
                        elif item == "Recording Dates":
                            excel_body[-1][item].append(str(sheet.cell(row=rnum,
                                                                       column=i+1).value.strftime("%Y-%m-%d")))
                        else:
                            excel_body[-1][item].append(sheet.cell(row=rnum,
                                                                   column=i+1).value)

        # get option dictionary
        # key: item name
        # value: item options
        excel_column = {}
        for item in excel_body:
            for name in excel_head:
                if name not in excel_column:
                    excel_column[name] = set(item[name])
                else:
                    excel_column[name] = excel_column[name].union(set(item[name]))

        self.content["head"] = excel_head
        self.content["body"] = excel_body
        self.content["column"] = excel_column

        self.set_configs()

    def get_options(self, item):
        return self.content["column"][item]

    # init configurations
    def set_configs(self, configs=None, switches=None):
        if configs:
            for i,j in configs:
                self.config[i] = j

        if switches is None:
            for item in self.content["head"]:
                self.switch[item] = [True, self.content["column"][item]]
        else:
            for i,j in switches:
                self.switch[i] = j

    # update output options
    def update_switch(self, item, enable, content="all", reverse=False):

        original = self.content["column"][item]

        if enable is False:
            self.switch[item] = [False, original]
            return

        if content is "all":
            self.switch[item] = [True, original]
            return

        if reverse:
            self.switch[item] = [True, set(original) - set(content)]
        else:
            self.switch[item] = [True, set(content)]

    # process switches and find its files to remove
    def run(self, output):
        lg.debug("generating to remove list...")
        remove_its = []

        # match information in child
        # get need to remove list
        for item in self.content["head"]:
            if self.switch[item][0]:
                nfilter = self.content["column"][item] - self.switch[item][1]
                for child in self.content["body"]:
                    if item == "Study Number":
                        if set(child["Study Number"]).issubset(nfilter):
                            remove_its.append(child["ITS File"])
                    else:
                        k = 0
                        for info in child[item]:
                            if set([info]).issubset(nfilter):
                                remove_its.append((child["ITS File"][k], item))
                            k += 1

        self.its_to_remove = remove_its

        if self.config['preliminary']:
            pfilename = os.path.dirname(output) + "/Special_Case.xlsx"
            self.save_preliminary(pfilename)

    def save_preliminary(self, filename):
        lg.debug("generating contents after filtering...")

        if os.path.isfile(filename):
            os.remove(filename)

        output = []
        head = [item for item in self.content['head'][1:] if self.switch[item][0]]

        # the first column is mandatory
        head = ["Study Number"] + head
        output.append(head)

        its_removed_list =[x[0] for x in self.its_to_remove]

        # record the filtered child information
        for child in self.content["body"]:
            # get the number of rows for this child
            max_len = len(child["ITS File"])

            for i in range(max_len):
                if child["ITS File"][i] not in its_removed_list:
                    new_output = []
                    new_output.append(child["Study Number"][0].lower())
                    for item in self.content["head"][1:]:
                        if self.switch[item][0]:
                            if len(child[item]) <= i or len(child[item][i]) == 0:
                                new_output.append(" ")
                            else:
                                new_output.append(child[item][i])
                    output.append(new_output)

        filter_output = []
        filter_output.append(["filtered files", "selected filter"])
        for i in self.its_to_remove:
            filter_output.append(i)

        mParser.excel_writer(filename, 'final', output)
        mParser.excel_writer(filename, 'filtered', filter_output)

