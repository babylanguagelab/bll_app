import openpyxl
import logging as lg
import ConfigParser as mParser

COMMENT_LIST = [
    'Study Number',
    'Recording',
    'Recording Dates',
    'Recording Type',
    'ITS File',
    'Recording Notes/Errors?',
    'Note in File',
    'Language Other than English',
    'Child Sick',
    'Short Recording',
    'Pause in Recording',
    'Unusual Behaviour',
    'LENA Device off Child',
    'Paired Recording (home AND homedaycare/daycare)',
    'Child Development Concern',
    'Clock Drift']


class CommentProcessor:
    def __init__(self):
        self.switches = [
            ['Recording', True],
            ['Recording Dates', True],
            ['Recording Type', True],
            ['ITS File', True],
            ['Recording Notes/Errors?', True],
            ['Language Other than English', True],
            ['Child Sick', True],
            ['Short Recording', True],
            ['Pause in Recording', True],
            ['Unusual Behaviour', True],
            ['LENA Device off Child', True],
            ['Paired Recording [home AND homedaycare/daycare]', True],
            ['Child Development Concern', True],
            ['Clock Drift', True]]

    def openCommentFile(self, filename):
        wb = openpyxl.load_workbook(filename)
        self.sheet = wb.get_sheet_by_name("Special Cases")
        self.heads = [x.value for x in self.sheet.rows[0]]
        self.content = [self.heads]

    def setSwitches(self, switches):
        for i in range(len(self.switches)):
            if self.switches[i][1] != switches[i]:
                self.switches[i][1] = switches[i]

    # process switches and filter content
    def run(self):
        switch_dict = dict(self.switches)
        switch_dict['Study Number'] = True
        new_heads = []

        for index in range(len(self.heads)):
            if self.heads[index] in switch_dict:
                if switch_dict[self.heads[index]]:
                    new_heads.append(index)

        for row_index in range(1, self.sheet.max_row):
            line = []
            for column_index in new_heads:
                line.append(self.sheet.cell(row=row_index+1,
                                            column=column_index+1).value)
            self.content.append(line)

    def saveResults(self, filename):
        mParser.excel_writer(filename, 'Special Cases', self.content)
