# csv, json, excel reader and writer
# author: zhangh15@myumanitoba.ca

import os.path
import csv
import json
import openpyxl


def json_reader(filename):
    content = {}
    with open(filename, 'rt') as fp:
        content = json.load(fp)
    return content

def json_writer(filename, data):
    with open(filename, 'wt') as fp:
        json.dump(data, fp)

def csv_reader(filename):
    content = []
    with open(filename, 'rt') as fp:
        reader = csv.reader(fp)
        for row in reader:
            content.append(row)
    return content

# write a list into a csv file
def csv_writer(filename, data):
    with open(filename, 'wt') as fp:
        writer = csv.writer(fp, quoting=csv.QUOTE_ALL)
        for row in data:
            writer.writerow(row)

# read CSV with specified header(column) names
def csv_dict_reader(filename, header):
    content = []
    with open(filename, 'rt') as fp:
        csv_content = csv.DictReader(fp)
        for row in csv_content:
            new_row = []
            for i in header:
                new_row.append(row[i])
            content.append(new_row)
    return content

# write a list of rows to excel
def excel_writer(filename, sheetname, data):
    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    sheet = wb.create_sheet(title=sheetname)

    for x in range(len(data)):
            for y in range(len(data[x])):
                sheet.cell(row=x+1, column=y+1).value = data[x][y]

    wb.save(filename)
